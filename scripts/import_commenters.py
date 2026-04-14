"""
import_commenters.py  [Phase 1 — 파일 기반 댓글 유저 import]
────────────────────────────────────────────────────────────
사용자가 외부 도구로 추출한 댓글 파일(xlsx/csv)에서
유저명을 읽어 Google Sheets 'ugc_users' 탭에 추가합니다.

지원 형식:
- xlsx: 첫 시트, A열에 username. (한국 도구 헤더: '아이디'도 자동 매핑)
- csv:  첫 컬럼이 username

실행:
  python scripts/import_commenters.py --file 인스타댓글.xlsx --source-url https://www.instagram.com/p/XXX/
"""

import os, sys, csv, argparse
from datetime import datetime, timezone
from dotenv import load_dotenv
import gspread
from google.oauth2.service_account import Credentials

load_dotenv()

SPREADSHEET_ID     = os.getenv("SPREADSHEET_ID")
GOOGLE_CREDENTIALS = os.getenv("GOOGLE_CREDENTIALS_PATH", "config/google_credentials.json")
SHEET_TAB_NAME     = "ugc_users"
MY_IG_ID           = "pitapat_prompt"

USERNAME_HEADERS = {"username", "user", "userid", "user_id", "아이디", "id"}


def read_xlsx(path: str) -> list[str]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    # username 컬럼 인덱스 찾기
    col_idx = 0
    for i, h in enumerate(header):
        if h in USERNAME_HEADERS:
            col_idx = i
            break
    return [str(r[col_idx]).strip() for r in rows[1:] if r and r[col_idx]]


def read_csv(path: str) -> list[str]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return []
    header = [c.strip().lower() for c in rows[0]]
    col_idx = 0
    for i, h in enumerate(header):
        if h in USERNAME_HEADERS:
            col_idx = i
            break
    return [r[col_idx].strip() for r in rows[1:] if r and r[col_idx]]


def get_sheet():
    creds = Credentials.from_service_account_file(
        GOOGLE_CREDENTIALS,
        scopes=["https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive"],
    )
    return gspread.authorize(creds).open_by_key(SPREADSHEET_ID).worksheet(SHEET_TAB_NAME)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True, help="댓글 추출 파일 (xlsx 또는 csv)")
    parser.add_argument("--source-url", default="", help="댓글 단 게시물 URL (메타용)")
    args = parser.parse_args()

    print("=" * 55)
    print("  UGC Monitor — Phase 1: 댓글 유저 Import")
    print("=" * 55)

    if not os.path.exists(args.file):
        print(f"❌ 파일 없음: {args.file}")
        sys.exit(1)

    ext = os.path.splitext(args.file)[1].lower()
    if ext == ".xlsx":
        usernames = read_xlsx(args.file)
    elif ext == ".csv":
        usernames = read_csv(args.file)
    else:
        print(f"❌ 지원하지 않는 형식: {ext} (xlsx, csv만 지원)")
        sys.exit(1)

    print(f"📄 파일에서 읽은 유저: {len(usernames)}명")

    # 본인 제외 + 중복 제거
    unique = set()
    cleaned = []
    for u in usernames:
        if not u:
            continue
        if u.lower() == MY_IG_ID.lower():
            continue
        if u.lower() in unique:
            continue
        unique.add(u.lower())
        cleaned.append(u)
    print(f"🧹 본인·중복 제거 후: {len(cleaned)}명")

    print("\nGoogle Sheets 연결 중...")
    sheet = get_sheet()
    existing = {r[0].strip().lower() for r in sheet.get_all_values()[1:] if r and r[0]}
    to_add = [u for u in cleaned if u.lower() not in existing]
    skipped = len(cleaned) - len(to_add)

    print(f"📋 기존 {len(existing)}명 | 신규 {len(cleaned)}명 → 추가 {len(to_add)}명 | 스킵 {skipped}명")

    if to_add:
        now  = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        # 컬럼 순서: username, instagram_id, added_at, last_checked, last_profile_url,
        #            has_story, is_ugc_detected, ugc_type, ugc_detected_at, source_post_url, notes
        rows = [[u, "", now, "", "", "", "", "none", "", args.source_url, ""] for u in to_add]
        sheet.append_rows(rows, value_input_option="RAW")
        print(f"✅ {len(to_add)}명 추가 완료!")
    else:
        print("→ 추가할 신규 유저 없음")

    print("\n" + "=" * 55)
    print(f"  완료 — 전체 {len(existing) + len(to_add)}명")
    print("=" * 55)


if __name__ == "__main__":
    main()
